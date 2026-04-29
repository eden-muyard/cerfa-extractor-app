import re
import unicodedata
from typing import Any

from openpyxl import load_workbook

from label_config import load_label_config

SIREN_REGEX = re.compile(r"\b(\d{3}\s?\d{3}\s?\d{3})\b")
AMOUNT_REGEX = re.compile(r"(-?\d[\d\s.,]*)")
YES_NO_REGEX = re.compile(r"\b(oui|non|yes|no)\b", re.IGNORECASE)
POLE_REGEX = re.compile(r"\b(energie|biotech|industrie|innovation|digitale)\b", re.IGNORECASE)
DUAL_CREDIT_REGEX = re.compile(r"\bcir\s*[\/\-\+]\s*cii\b|\bcii\s*[\/\-\+]\s*cir\b", re.IGNORECASE)
YEAR_REGEX = re.compile(r"\b(20\d{2})\b")
MAX_SCAN_ROWS = 2500
MAX_SCAN_COLS = 120


def normalize_amount(raw: str) -> str:
    value = raw.replace(" ", "").replace("\u00a0", "")
    if value.count(",") == 1 and value.count(".") == 0:
        value = value.replace(",", ".")
    value = value.replace(",", "")
    return value


def parse_with_label_nearby(next_value: Any) -> str | None:
    if next_value is None:
        return None
    if isinstance(next_value, (int, float)):
        return str(next_value)
    cleaned = str(next_value).strip()
    return cleaned if cleaned else None


def compile_patterns(raw_patterns: list[str]) -> list[re.Pattern]:
    return [re.compile(pattern, re.IGNORECASE) for pattern in raw_patterns]


def normalize_sheet_name(name: str) -> str:
    no_accents = unicodedata.normalize("NFKD", name).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", no_accents.strip().lower())


def canonical_sheet_name(raw_name: str) -> str | None:
    normalized = normalize_sheet_name(raw_name)
    if "synthese" in normalized:
        return "synthese"
    if "parametr" in normalized and "chiffrage" in normalized:
        return "parametres_chiffrage"
    if "rep cout" in normalized or ("rep" in normalized and "cout" in normalized):
        return "rep_cout"
    if "2069" in normalized and "a" in normalized:
        return "2069_a"
    return None


def normalize_yes_no(raw: str) -> str:
    lowered = raw.strip().lower()
    if lowered in {"oui", "yes"}:
        return "Oui"
    if lowered in {"non", "no"}:
        return "Non"
    return raw


def normalize_pole(raw: str) -> str:
    match = POLE_REGEX.search(raw)
    if not match:
        return raw
    value = match.group(1).lower()
    return value.capitalize()


def normalize_honoraires_type(raw: str) -> str:
    text = normalize_text(raw)
    if "palier" in text:
        return "Fixe a palier"
    if "fixe" in text:
        return "Fixe"
    if "%" in str(raw) or "pourcentage" in text or "percent" in text:
        return "%age"
    return str(raw).strip()


def normalize_text(raw: Any) -> str:
    if raw is None:
        return ""
    text = str(raw).strip()
    no_accents = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", no_accents.lower())


def get_candidate_values(row_values: tuple[Any, ...], index: int) -> list[Any]:
    candidates: list[Any] = []
    for offset in range(1, 13):
        target_index = index + offset
        if target_index >= len(row_values):
            break
        candidate = row_values[target_index]
        if candidate is None:
            continue
        candidate_text = str(candidate).strip()
        if candidate_text:
            candidates.append(candidate)
    return candidates


def extract_number(raw: Any) -> str | None:
    text = str(raw).strip()
    amount_match = AMOUNT_REGEX.search(text)
    if not amount_match:
        return None
    number = normalize_amount(amount_match.group(1))
    try:
        as_float = float(number)
    except ValueError:
        return None
    # Ignore year-only values mistakenly captured as amounts.
    if as_float.is_integer() and 1900 <= int(as_float) <= 2100:
        return None
    return number


def extract_credit_choice(raw: Any) -> str | None:
    text = normalize_text(raw)
    if not text:
        return None
    if re.search(r"\bcico\b", text):
        return "CICO"
    if re.search(r"\bcic\b", text):
        return "CIC"
    if DUAL_CREDIT_REGEX.search(text):
        return "CIR/CII"
    has_cir = re.search(r"\bcir\b", text) is not None
    has_cii = re.search(r"\bcii\b", text) is not None
    if has_cir and has_cii:
        return "CIR/CII"
    if has_cir:
        return "CIR"
    if has_cii:
        return "CII"
    return None


def extract_year(raw: Any) -> str | None:
    text = normalize_text(raw)
    if not text:
        return None
    matches = YEAR_REGEX.findall(text)
    valid = [year for year in matches if 2018 <= int(year) <= 2035]
    if not valid:
        return None
    return str(max(int(year) for year in valid))


def extract_text_candidate(raw: Any) -> str | None:
    text = str(raw).strip()
    if not text:
        return None
    lowered = normalize_text(text)
    if lowered in {"none", "null", "nan"}:
        return None
    if lowered.startswith("ligne "):
        return None
    if re.fullmatch(r"[-\d\s.,]+", text):
        return None
    return text


def find_header_keyword_above(
    rows: list[tuple[Any, ...]], row_idx: int, col_idx: int, keyword: str
) -> bool:
    for up in range(1, 5):
        header_row_idx = row_idx - up
        if header_row_idx < 0:
            break
        header_row = rows[header_row_idx]
        if col_idx >= len(header_row):
            continue
        header_value = header_row[col_idx]
        if header_value is None:
            continue
        if keyword in normalize_text(header_value):
            return True
    return False


def extract_credit_amounts_from_synthese(sheet) -> dict[str, str]:
    found = {"montant_credit_impot_cir": "", "montant_credit_impot_cii": "", "montant_cic": ""}
    max_row = min(sheet.max_row or 400, 400)
    max_col = min(sheet.max_column or 40, 40)
    rows = list(
        sheet.iter_rows(
            min_row=1,
            max_row=max_row,
            min_col=1,
            max_col=max_col,
            values_only=True,
        )
    )

    for row_idx, row in enumerate(rows):
        normalized_cells = [normalize_text(value) for value in row]
        if not any("montant du credit d'impot" in cell or "montant du credit d impot" in cell for cell in normalized_cells):
            continue

        for col_idx, value in enumerate(row):
            if not found["montant_credit_impot_cir"] and find_header_keyword_above(rows, row_idx, col_idx, "cir"):
                amount = extract_number(value)
                if amount:
                    found["montant_credit_impot_cir"] = amount
            if not found["montant_credit_impot_cii"] and find_header_keyword_above(rows, row_idx, col_idx, "cii"):
                amount = extract_number(value)
                if amount:
                    found["montant_credit_impot_cii"] = amount
            if not found["montant_cic"] and find_header_keyword_above(rows, row_idx, col_idx, "cic"):
                amount = extract_number(value)
                if amount:
                    found["montant_cic"] = amount

        # Fallback: many matrices place these in C/D on the same row.
        if not found["montant_credit_impot_cir"] and len(row) > 2:
            amount = extract_number(row[2])
            if amount:
                found["montant_credit_impot_cir"] = amount
        if not found["montant_credit_impot_cii"] and len(row) > 3:
            amount = extract_number(row[3])
            if amount:
                found["montant_credit_impot_cii"] = amount
        if not found["montant_cic"] and len(row) > 4:
            amount = extract_number(row[4])
            if amount:
                found["montant_cic"] = amount

        if found["montant_credit_impot_cir"] and found["montant_credit_impot_cii"] and found["montant_cic"]:
            break

    return found


def parse_number_to_float(value: str | None) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except ValueError:
        return None


def pick_value_by_credit(cir_value: str | None, cii_value: str | None, credit_choice: str | None) -> str | None:
    cir_float = parse_number_to_float(cir_value)
    cii_float = parse_number_to_float(cii_value)
    normalized_credit = (credit_choice or "").upper()

    if normalized_credit == "CII":
        if cii_value is not None:
            return cii_value
        return cir_value
    if normalized_credit == "CIR":
        if cir_value is not None:
            return cir_value
        return cii_value
    if normalized_credit in {"CIR/CII", "CIC", "CICO"}:
        total = (cir_float or 0.0) + (cii_float or 0.0)
        return format_total_number(total)

    if cii_float and cii_float != 0:
        return cii_value
    if cir_float and cir_float != 0:
        return cir_value
    return cii_value if cii_value is not None else cir_value


def find_cir_cii_columns(rows: list[tuple[Any, ...]]) -> tuple[int | None, int | None]:
    for row in rows[:40]:
        for col_idx, value in enumerate(row):
            text = normalize_text(value)
            if not text:
                continue
            if re.fullmatch(r"cir", text):
                cir_col = col_idx
                cii_col = None
                for idx2, value2 in enumerate(row):
                    text2 = normalize_text(value2)
                    if re.fullmatch(r"cii", text2):
                        cii_col = idx2
                        break
                if cii_col is not None:
                    return cir_col, cii_col
    return None, None


def row_matches_tokens(normalized_row_text: str, tokens: list[str]) -> bool:
    return all(token in normalized_row_text for token in tokens)


def extract_number_from_column(row: tuple[Any, ...], col_idx: int | None) -> str | None:
    if col_idx is None or col_idx >= len(row):
        return None
    return extract_number(row[col_idx])


def extract_synthese_depenses_by_columns(sheet, credit_choice: str | None) -> dict[str, str]:
    fields = {
        "dotations_amortissements": ["dotations", "amort"],
        "depenses_personnel": ["depenses", "personnel"],
        "depenses_veille": ["depenses", "veille"],
        "depenses_brevets": ["depenses", "brevets"],
        "depenses_fonctionnement": ["depenses", "fonctionnement"],
        "subventions_avances_remboursables": ["subventions", "avances", "remboursables"],
    }
    result = {key: "" for key in fields}
    max_row = min(sheet.max_row or 500, 500)
    max_col = min(sheet.max_column or 60, 60)
    rows = list(
        sheet.iter_rows(
            min_row=1,
            max_row=max_row,
            min_col=1,
            max_col=max_col,
            values_only=True,
        )
    )
    cir_col, cii_col = find_cir_cii_columns(rows)
    if cir_col is None and cii_col is None:
        return result

    for row in rows:
        normalized_row_text = " ".join(
            normalize_text(value) for value in row if value is not None and str(value).strip()
        )
        if not normalized_row_text:
            continue

        for field, tokens in fields.items():
            if result[field]:
                continue
            if field == "depenses_personnel" and "jeunes docteurs" in normalized_row_text:
                continue
            if row_matches_tokens(normalized_row_text, tokens):
                cir_value = extract_number_from_column(row, cir_col)
                cii_value = extract_number_from_column(row, cii_col)
                selected = pick_value_by_credit(cir_value, cii_value, credit_choice)
                if selected is not None:
                    result[field] = selected
    return result


def format_total_number(total: float) -> str:
    if float(total).is_integer():
        return str(int(total))
    return f"{total:.2f}".rstrip("0").rstrip(".")


def extract_prestataires_total_from_synthese(sheet, credit_choice: str | None) -> str:
    targets = [
        ["prestataires", "externes", "francais"],
        ["prestataires", "externes", "communautaires"],
        ["prestataires", "francais", "lien", "dependance"],
        ["prestataires", "communautaires", "lien", "dependance"],
    ]
    max_row = min(sheet.max_row or 500, 500)
    max_col = min(sheet.max_column or 60, 60)
    rows = list(
        sheet.iter_rows(
            min_row=1,
            max_row=max_row,
            min_col=1,
            max_col=max_col,
            values_only=True,
        )
    )
    cir_col, cii_col = find_cir_cii_columns(rows)
    seen = set()
    total = 0.0

    for row in rows:
        row_normalized = " ".join(
            normalize_text(value) for value in row if value is not None and str(value).strip()
        )
        if not row_normalized:
            continue

        matched_idx = None
        for idx, tokens in enumerate(targets):
            if idx in seen:
                continue
            if all(token in row_normalized for token in tokens):
                matched_idx = idx
                break
        if matched_idx is None:
            continue

        cir_value = extract_number_from_column(row, cir_col)
        cii_value = extract_number_from_column(row, cii_col)
        selected = pick_value_by_credit(cir_value, cii_value, credit_choice)
        selected_float = parse_number_to_float(selected)
        if selected_float is None:
            continue

        seen.add(matched_idx)
        total += selected_float
        if len(seen) == len(targets):
            break

    if total == 0.0 and len(seen) == 0:
        return ""
    return format_total_number(total)


def count_project_columns_from_rep_cout(sheet) -> dict[str, str]:
    def is_value_like(raw: Any) -> bool:
        if raw is None:
            return False
        text = str(raw).strip()
        if not text or text in {"-", "--", "---"}:
            return False
        number = extract_number(raw)
        parsed = parse_number_to_float(number)
        return parsed is not None and abs(parsed) > 0

    max_row = min(sheet.max_row or 240, 240)
    max_col = min(sheet.max_column or 220, 220)
    rows = list(
        sheet.iter_rows(
            min_row=1,
            max_row=max_row,
            min_col=1,
            max_col=max_col,
            values_only=True,
        )
    )

    def find_section_anchor(section_key: str) -> int | None:
        for row_idx, row in enumerate(rows):
            for value in row:
                text = normalize_text(value)
                if section_key == "cir" and re.search(r"credit d'impot recherche|\bcir\b", text):
                    return row_idx
                if section_key == "cii" and re.search(r"credit d'impot innovation|\bcii\b", text):
                    return row_idx
        return None

    def count_projects_in_section(section_key: str, start_row: int, end_row: int) -> int:
        candidate_cols: set[int] = set()
        for row_idx in range(start_row, min(end_row + 1, start_row + 6)):
            row = rows[row_idx]
            for col_idx, value in enumerate(row):
                text = normalize_text(value)
                if not text or col_idx == 0:
                    continue
                if "total" in text or "non rattache" in text:
                    continue
                if "projet" in text or re.fullmatch(r"cir", text) or re.fullmatch(r"cii", text):
                    candidate_cols.add(col_idx)

        if not candidate_cols:
            max_len = max((len(rows[r]) for r in range(start_row, end_row + 1)), default=0)
            for col_idx in range(1, max_len):
                header_text = " ".join(
                    normalize_text(rows[r][col_idx])
                    for r in range(start_row, min(end_row + 1, start_row + 6))
                    if col_idx < len(rows[r]) and rows[r][col_idx] is not None
                )
                if "total" in header_text or "non rattache" in header_text:
                    continue
                if "projet" in header_text or section_key in header_text:
                    candidate_cols.add(col_idx)

        value_rows = range(min(end_row, start_row + 3), end_row + 1)
        count = 0
        for col_idx in sorted(candidate_cols):
            has_value = False
            for row_idx in value_rows:
                row = rows[row_idx]
                if col_idx >= len(row):
                    continue
                left_label = normalize_text(row[0]) if row and row[0] is not None else ""
                if "total" in left_label:
                    continue
                if is_value_like(row[col_idx]):
                    has_value = True
                    break
            if has_value:
                count += 1
        return count

    cir_anchor = find_section_anchor("cir")
    cii_anchor = find_section_anchor("cii")
    cir_count = 0
    cii_count = 0
    if cir_anchor is not None:
        cir_end = (cii_anchor - 1) if cii_anchor is not None and cii_anchor > cir_anchor else min(len(rows) - 1, cir_anchor + 70)
        cir_count = count_projects_in_section("cir", cir_anchor, cir_end)
    if cii_anchor is not None:
        cii_end = min(len(rows) - 1, cii_anchor + 70)
        cii_count = count_projects_in_section("cii", cii_anchor, cii_end)

    return {
        "nombre_projets_cir": str(cir_count),
        "nombre_projets_cii": str(cii_count),
    }


def extract_honoraires_n_1_from_parametres(sheet, annee_valorisation: str | None) -> str:
    def get_year(value: Any) -> int | None:
        text = normalize_text(value)
        if not text:
            return None
        match = YEAR_REGEX.search(text)
        if not match:
            return None
        year = int(match.group(1))
        return year if 1990 <= year <= 2100 else None

    def is_hono_label(value: Any) -> bool:
        text = normalize_text(value)
        return "hono" in text or "honoraire" in text

    def number_at(rows: list[list[Any]], row_idx: int, col_idx: int) -> str:
        if row_idx < 0 or col_idx < 0:
            return ""
        if row_idx >= len(rows):
            return ""
        row = rows[row_idx]
        if col_idx >= len(row):
            return ""
        extracted = extract_number(row[col_idx])
        return extracted or ""

    target_year: int | None = None
    if annee_valorisation and annee_valorisation.isdigit():
        target_year = int(annee_valorisation) - 1
    if target_year is None:
        return ""

    max_row = min(sheet.max_row or 500, 500)
    max_col = min(sheet.max_column or 120, 120)
    rows = [
        list(row)
        for row in sheet.iter_rows(
            min_row=1,
            max_row=max_row,
            min_col=1,
            max_col=max_col,
            values_only=True,
        )
    ]

    target_year_positions: list[tuple[int, int]] = []
    hono_positions: list[tuple[int, int]] = []
    for r_idx, row in enumerate(rows):
        for c_idx, value in enumerate(row):
            year = get_year(value)
            if year == target_year:
                target_year_positions.append((r_idx, c_idx))
            if is_hono_label(value):
                hono_positions.append((r_idx, c_idx))

    for y_row, y_col in target_year_positions:
        for h_row, h_col in hono_positions:
            if abs(y_row - h_row) <= 8:
                amount = number_at(rows, h_row, y_col)
                if amount:
                    return amount
                for delta in (-2, -1, 1, 2):
                    amount = number_at(rows, h_row, y_col + delta)
                    if amount:
                        return amount
            if abs(y_col - h_col) <= 8:
                amount = number_at(rows, y_row, h_col)
                if amount:
                    return amount
                for delta in (-2, -1, 1, 2):
                    amount = number_at(rows, y_row + delta, h_col)
                    if amount:
                        return amount

    # Do not fallback to another year when requested target year exists but no aligned value found.
    return ""


def extract_line_amount_from_2069_row(row: tuple[Any, ...]) -> float | None:
    numbers: list[float] = []
    for value in row:
        extracted = extract_number(value)
        as_float = parse_number_to_float(extracted)
        if as_float is not None:
            numbers.append(as_float)
    if not numbers:
        return None
    return numbers[-1]


def extract_2069_line_totals(sheet) -> dict[str, str]:
    internal_lines = {7, 8, 9, 10, 11, 12, 13}
    external_lines = {14, 21}
    internal_sum = 0.0
    external_sum = 0.0
    found_internal: set[int] = set()
    found_external: set[int] = set()
    max_row = min(sheet.max_row or 500, 500)
    max_col = min(sheet.max_column or 80, 80)
    for row in sheet.iter_rows(
        min_row=1,
        max_row=max_row,
        min_col=1,
        max_col=max_col,
        values_only=True,
    ):
        row_text = " ".join(
            normalize_text(value) for value in row if value is not None and str(value).strip()
        )
        if "ligne" not in row_text:
            continue
        for line_no in internal_lines.union(external_lines):
            if line_no in found_internal or line_no in found_external:
                continue
            if not re.search(rf"\bligne\s*{line_no}\b", row_text):
                continue
            amount = extract_line_amount_from_2069_row(row)
            if amount is None:
                continue
            if line_no in internal_lines:
                internal_sum += amount
                found_internal.add(line_no)
            if line_no in external_lines:
                external_sum += amount
                found_external.add(line_no)
    return {
        "depenses_internes": format_total_number(internal_sum) if found_internal else "",
        "depenses_externes": format_total_number(external_sum) if found_external else "",
    }


def extract_fields_from_workbook(file_path: str) -> dict[str, str]:
    configured_fields = load_label_config()
    field_patterns = {
        field["key"]: compile_patterns(field.get("patterns", []))
        for field in configured_fields
    }
    field_types = {field["key"]: field.get("value_type", "text") for field in configured_fields}
    field_source_tabs = {
        field["key"]: set(field.get("source_tabs", []))
        for field in configured_fields
    }

    workbook = load_workbook(file_path, data_only=True, read_only=True)
    data: dict[str, str] = {field["key"]: "" for field in configured_fields}

    target_worksheets = [
        sheet for sheet in workbook.worksheets if canonical_sheet_name(sheet.title)
    ]

    for sheet in target_worksheets:
        sheet_key = canonical_sheet_name(sheet.title)
        if not sheet_key:
            continue
        credit_choice = data.get("credit_parametres", "")
        if sheet_key == "rep_cout":
            project_counts = count_project_columns_from_rep_cout(sheet)
            for key, value in project_counts.items():
                if key in data and value:
                    data[key] = value
        if sheet_key == "parametres_chiffrage":
            honoraires_n_1 = extract_honoraires_n_1_from_parametres(sheet, data.get("annee_valorisation"))
            if "honoraires_n_1" in data and honoraires_n_1:
                data["honoraires_n_1"] = honoraires_n_1
        if sheet_key == "synthese":
            credit_values = extract_credit_amounts_from_synthese(sheet)
            for key, value in credit_values.items():
                if key in data and value and not data[key]:
                    data[key] = value
            depenses_from_columns = extract_synthese_depenses_by_columns(sheet, credit_choice)
            for key, value in depenses_from_columns.items():
                if key in data and value:
                    current_float = parse_number_to_float(data.get(key))
                    new_float = parse_number_to_float(value)
                    if current_float is None or (new_float is not None and abs(new_float) > abs(current_float)):
                        data[key] = value
            if "depenses_prestataires_externes" in data and not data["depenses_prestataires_externes"]:
                prestataires_total = extract_prestataires_total_from_synthese(sheet, credit_choice)
                if prestataires_total:
                    data["depenses_prestataires_externes"] = prestataires_total
        if sheet_key == "2069_a":
            totals_2069 = extract_2069_line_totals(sheet)
            for key, value in totals_2069.items():
                if key in data and value:
                    data[key] = value
        max_row = min(sheet.max_row or MAX_SCAN_ROWS, MAX_SCAN_ROWS)
        max_col = min(sheet.max_column or MAX_SCAN_COLS, MAX_SCAN_COLS)
        for row in sheet.iter_rows(
            min_row=1,
            max_row=max_row,
            min_col=1,
            max_col=max_col,
            values_only=True,
        ):
            for index, value in enumerate(row):
                if value is None:
                    continue
                cell_text = str(value).strip()
                if not cell_text:
                    continue
                normalized_cell_text = normalize_text(value)
                candidate_values = get_candidate_values(row, index)
                if not candidate_values and index + 1 < len(row):
                    candidate_values = [row[index + 1]]

                for field, patterns in field_patterns.items():
                    if data[field]:
                        continue
                    source_tabs = field_source_tabs.get(field, set())
                    if source_tabs and sheet_key not in source_tabs:
                        continue
                    for pattern in patterns:
                        if pattern.search(cell_text) or pattern.search(normalized_cell_text):
                            value_type = field_types.get(field, "text")
                            extracted: str | None = None
                            if value_type == "siren":
                                for candidate in candidate_values + [cell_text]:
                                    siren_match = SIREN_REGEX.search(str(candidate))
                                    if siren_match:
                                        extracted = siren_match.group(1).replace(" ", "")
                                        break
                            elif value_type == "number":
                                for candidate in candidate_values:
                                    extracted = extract_number(candidate)
                                    if extracted:
                                        break
                                if not extracted:
                                    extracted = extract_number(cell_text)
                            elif value_type == "yes_no":
                                for candidate in candidate_values + [cell_text]:
                                    yes_no_match = YES_NO_REGEX.search(str(candidate))
                                    if yes_no_match:
                                        extracted = normalize_yes_no(yes_no_match.group(1))
                                        break
                            elif value_type == "pole":
                                for candidate in candidate_values + [cell_text]:
                                    pole_match = POLE_REGEX.search(str(candidate))
                                    if pole_match:
                                        extracted = normalize_pole(pole_match.group(1))
                                        break
                            elif value_type == "credit_choice":
                                for candidate in candidate_values + [cell_text]:
                                    extracted = extract_credit_choice(candidate)
                                    if extracted:
                                        break
                            elif value_type == "year":
                                for candidate in candidate_values + [cell_text]:
                                    extracted = extract_year(candidate)
                                    if extracted:
                                        break
                            elif value_type == "honoraires_type":
                                for candidate in candidate_values + [cell_text]:
                                    candidate_text = str(candidate).strip()
                                    normalized = normalize_text(candidate_text)
                                    if (
                                        "palier" in normalized
                                        or "fixe" in normalized
                                        or "%" in candidate_text
                                        or "pourcentage" in normalized
                                    ):
                                        extracted = normalize_honoraires_type(candidate_text)
                                        break
                            if not extracted and value_type == "text":
                                for candidate in candidate_values:
                                    extracted = extract_text_candidate(candidate)
                                    if extracted:
                                        break
                            if value_type == "yes_no" and extracted:
                                yes_no_match = YES_NO_REGEX.search(extracted)
                                if yes_no_match:
                                    extracted = normalize_yes_no(yes_no_match.group(1))
                            if value_type == "pole" and extracted:
                                extracted = normalize_pole(extracted)
                            if extracted:
                                data[field] = extracted
                            break
            if all(data.values()):
                break
        if all(data.values()):
            break

    return data